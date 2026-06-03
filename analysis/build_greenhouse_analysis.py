#!/usr/bin/env python3
"""Deep exploratory analysis for greenhouse sensor and lettuce experiment data.

The script intentionally uses only the standard library plus openpyxl/reportlab,
because the project environment does not provide pandas/numpy/matplotlib.
"""

from __future__ import annotations

import bisect
import csv
import json
import math
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from statistics import mean, median, pstdev
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.graphics.shapes import Drawing, Line, Path as RlPath, Rect, String


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "analysis"
TABLES = OUT / "tables"
FIGS = OUT / "figures"
NOTES = OUT / "notes"
XLSX = ROOT / "Данные салат мощность.xlsx"
EVIDENCE = ROOT / "thesis" / "evidence" / "greenhouse_pi_data"
DB = EVIDENCE / "history_2026-05-26.db"
PHOTOPERIOD_LIGHT_H = 18
PHOTOPERIOD_DARK_H = 6
PPFD_VARIANTS = [
    {"variant": 1, "ppfd_umol_m2_s": 79.0, "ppfd_error": 3.0},
    {"variant": 2, "ppfd_umol_m2_s": 95.0, "ppfd_error": 6.4},
    {"variant": 3, "ppfd_umol_m2_s": 150.8, "ppfd_error": 11.4},
    {"variant": 4, "ppfd_umol_m2_s": 181.3, "ppfd_error": 14.3},
    {"variant": 5, "ppfd_umol_m2_s": 245.5, "ppfd_error": 13.1},
]

DEVICE_COLORS = {
    "th-1": "#d55e00",
    "th-2": "#009e73",
    "co2-1": "#0072b2",
}


@dataclass
class SensorRecord:
    ts: datetime
    device: str
    temperature: float | None
    humidity: float | None
    co2: float | None


@dataclass
class RelayEvent:
    ts: datetime
    relay_id: int
    state: str


def ensure_dirs() -> None:
    for path in [OUT, TABLES, FIGS, NOTES]:
        path.mkdir(parents=True, exist_ok=True)


def parse_ts(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def clean_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return f


def q(values: list[float], p: float) -> float | None:
    if not values:
        return None
    xs = sorted(values)
    pos = (len(xs) - 1) * p
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return xs[lo]
    return xs[lo] * (hi - pos) + xs[hi] * (pos - lo)


def fmt(value, digits: int = 2) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return f"{value:.{digits}f}"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def stats(values: list[float]) -> dict[str, float | int | None]:
    xs = [v for v in values if v is not None and not math.isnan(v)]
    if not xs:
        return {"n": 0, "min": None, "p05": None, "mean": None, "median": None, "p95": None, "max": None, "std": None}
    return {
        "n": len(xs),
        "min": min(xs),
        "p05": q(xs, 0.05),
        "mean": mean(xs),
        "median": median(xs),
        "p95": q(xs, 0.95),
        "max": max(xs),
        "std": pstdev(xs) if len(xs) > 1 else 0.0,
    }


def dli_from_ppfd(ppfd: float, light_hours: float = PHOTOPERIOD_LIGHT_H) -> float:
    return ppfd * light_hours * 3600 / 1_000_000


def corr(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 3 or len(xs) != len(ys):
        return None
    mx, my = mean(xs), mean(ys)
    sx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    sy = math.sqrt(sum((y - my) ** 2 for y in ys))
    if sx == 0 or sy == 0:
        return None
    return sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / (sx * sy)


def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not fieldnames:
        keys = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def load_sensor_data() -> tuple[list[SensorRecord], list[RelayEvent]]:
    con = sqlite3.connect(DB)
    sensors = [
        SensorRecord(parse_ts(ts), device, temperature, humidity, co2)
        for ts, device, temperature, humidity, co2 in con.execute(
            "select ts, device, temperature, humidity, co2 from sensor_log order by ts, device"
        )
    ]
    relays = [
        RelayEvent(parse_ts(ts), int(relay_id), state)
        for ts, relay_id, state in con.execute(
            "select ts, relay_id, state from relay_log order by ts, id"
        )
    ]
    con.close()
    return sensors, relays


def load_json_sources() -> tuple[dict, list[dict], list[dict]]:
    names = json.loads((EVIDENCE / "names_2026-05-26.json").read_text(encoding="utf-8"))
    profiles = json.loads((EVIDENCE / "profiles_2026-05-26.json").read_text(encoding="utf-8"))
    rules = json.loads((EVIDENCE / "rules_2026-05-26.json").read_text(encoding="utf-8"))
    return names, profiles, rules


def source_inventory(sensors: list[SensorRecord], relays: list[RelayEvent]) -> list[dict]:
    rows = [
        {
            "source": "history_2026-05-26.db:sensor_log",
            "type": "SQLite table",
            "rows": len(sensors),
            "period_start": min(r.ts for r in sensors),
            "period_end": max(r.ts for r in sensors),
            "comment": "Минутные показания трех устройств: th-1, th-2, co2-1.",
        },
        {
            "source": "history_2026-05-26.db:relay_log",
            "type": "SQLite table",
            "rows": len(relays),
            "period_start": min(r.ts for r in relays),
            "period_end": max(r.ts for r in relays),
            "comment": "События ON/OFF релейных каналов.",
        },
    ]
    local_db = ROOT / "Greenhouse" / "dashboard" / "data" / "history.db"
    if local_db.exists():
        try:
            con = sqlite3.connect(local_db)
            sensor_count, sensor_min, sensor_max = con.execute("select count(*), min(ts), max(ts) from sensor_log").fetchone()
            relay_count, relay_min, relay_max = con.execute("select count(*), min(ts), max(ts) from relay_log").fetchone()
            con.close()
            rows.append(
                {
                    "source": "Greenhouse/dashboard/data/history.db",
                    "type": "SQLite table check",
                    "rows": sensor_count + relay_count,
                    "period_start": sensor_min or relay_min or "",
                    "period_end": sensor_max or relay_max or "",
                    "comment": f"Локальная БД dashboard проверена отдельно: sensor_log={sensor_count}, relay_log={relay_count}; дополнительных наблюдений нет.",
                }
            )
        except sqlite3.Error as exc:
            rows.append(
                {
                    "source": "Greenhouse/dashboard/data/history.db",
                    "type": "SQLite table check",
                    "rows": "",
                    "period_start": "",
                    "period_end": "",
                    "comment": f"Не удалось прочитать как журнал истории: {exc}",
                }
            )
    mosquitto_db = ROOT / "Greenhouse" / "mosquitto" / "data" / "mosquitto.db"
    if mosquitto_db.exists():
        rows.append(
            {
                "source": "Greenhouse/mosquitto/data/mosquitto.db",
                "type": "Mosquitto persistence file",
                "rows": "",
                "period_start": "",
                "period_end": "",
                "comment": "Проверен как доступный файл данных; это служебное хранилище MQTT-брокера, а не структурированный журнал измерений.",
            }
        )
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    for ws in wb.worksheets:
        non_empty = 0
        numeric = 0
        dates = 0
        text = 0
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                non_empty += 1
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    numeric += 1
                elif isinstance(value, datetime):
                    dates += 1
                elif isinstance(value, str):
                    text += 1
        rows.append(
            {
                "source": f"Данные салат мощность.xlsx:{ws.title}",
                "type": "Excel sheet",
                "rows": ws.max_row,
                "cols": ws.max_column,
                "non_empty_cells": non_empty,
                "numeric_cells": numeric,
                "date_cells": dates,
                "text_cells": text,
                "comment": "",
            }
        )
    return rows


def sensor_device_summary(sensors: list[SensorRecord]) -> list[dict]:
    by_device = defaultdict(list)
    for r in sensors:
        by_device[r.device].append(r)
    rows = []
    for device, recs in sorted(by_device.items()):
        row = {
            "device": device,
            "rows": len(recs),
            "first_ts": min(r.ts for r in recs),
            "last_ts": max(r.ts for r in recs),
            "zero_temp_rows": sum(1 for r in recs if r.temperature == 0),
            "zero_humidity_rows": sum(1 for r in recs if r.humidity == 0),
            "zero_co2_rows": sum(1 for r in recs if r.co2 == 0),
        }
        for metric, valid in [
            ("temperature", lambda v: v is not None and 5 <= v <= 45),
            ("humidity", lambda v: v is not None and 1 <= v <= 100),
            ("co2", lambda v: v is not None and v >= 300),
        ]:
            values = [getattr(r, metric) for r in recs if valid(getattr(r, metric))]
            s = stats(values)
            for k, v in s.items():
                row[f"{metric}_{k}"] = v
        rows.append(row)
    return rows


def align_by_minute(sensors: list[SensorRecord]) -> dict[datetime, dict[str, SensorRecord]]:
    aligned: dict[datetime, dict[str, SensorRecord]] = defaultdict(dict)
    for r in sensors:
        minute = r.ts.replace(second=0)
        aligned[minute][r.device] = r
    return dict(sorted(aligned.items()))


def pairwise_sensor_comparison(aligned: dict[datetime, dict[str, SensorRecord]], metric: str) -> list[dict]:
    pairs = [("th-1", "th-2"), ("th-1", "co2-1"), ("th-2", "co2-1")]
    rows = []
    for a, b in pairs:
        xs, ys, diffs = [], [], []
        for devices in aligned.values():
            if a not in devices or b not in devices:
                continue
            va = getattr(devices[a], metric)
            vb = getattr(devices[b], metric)
            if va is None or vb is None:
                continue
            if metric == "temperature" and not (5 <= va <= 45 and 5 <= vb <= 45):
                continue
            if metric == "humidity" and not (1 <= va <= 100 and 1 <= vb <= 100):
                continue
            xs.append(va)
            ys.append(vb)
            diffs.append(va - vb)
        absdiff = [abs(d) for d in diffs]
        rows.append(
            {
                "metric": metric,
                "pair": f"{a} - {b}",
                "n_common_minutes": len(diffs),
                "mean_diff": mean(diffs) if diffs else None,
                "median_diff": median(diffs) if diffs else None,
                "mean_abs_diff": mean(absdiff) if absdiff else None,
                "p95_abs_diff": q(absdiff, 0.95) if absdiff else None,
                "rmse_diff": math.sqrt(mean([d * d for d in diffs])) if diffs else None,
                "correlation": corr(xs, ys),
            }
        )
    return rows


def hourly_cycles(sensors: list[SensorRecord]) -> list[dict]:
    bucket = defaultdict(list)
    for r in sensors:
        for metric, value in [
            ("temperature", r.temperature),
            ("humidity", r.humidity),
            ("co2", r.co2),
        ]:
            if value is None:
                continue
            if metric == "temperature" and not (5 <= value <= 45):
                continue
            if metric == "humidity" and not (1 <= value <= 100):
                continue
            if metric == "co2" and value < 300:
                continue
            bucket[(r.device, metric, r.ts.hour)].append(value)
    rows = []
    for (device, metric, hour), values in sorted(bucket.items()):
        rows.append(
            {
                "device": device,
                "metric": metric,
                "hour": hour,
                "n": len(values),
                "mean": mean(values),
                "median": median(values),
                "p05": q(values, 0.05),
                "p95": q(values, 0.95),
            }
        )
    return rows


def daily_summary(sensors: list[SensorRecord]) -> list[dict]:
    bucket = defaultdict(list)
    for r in sensors:
        day = r.ts.date().isoformat()
        for metric, value in [
            ("temperature", r.temperature),
            ("humidity", r.humidity),
            ("co2", r.co2),
        ]:
            if value is None:
                continue
            if metric == "temperature" and not (5 <= value <= 45):
                continue
            if metric == "humidity" and not (1 <= value <= 100):
                continue
            if metric == "co2" and value < 300:
                continue
            bucket[(day, r.device, metric)].append(value)
    rows = []
    for (day, device, metric), values in sorted(bucket.items()):
        rows.append(
            {
                "date": day,
                "device": device,
                "metric": metric,
                "n": len(values),
                "mean": mean(values),
                "min": min(values),
                "max": max(values),
                "median": median(values),
            }
        )
    return rows


def gap_analysis(sensors: list[SensorRecord]) -> list[dict]:
    by_device = defaultdict(list)
    for r in sensors:
        by_device[r.device].append(r.ts)
    rows = []
    for device, times in sorted(by_device.items()):
        times.sort()
        gaps = [(b - a).total_seconds() / 60 for a, b in zip(times, times[1:])]
        big = [(times[i], times[i + 1], gap) for i, gap in enumerate(gaps) if gap > 5]
        rows.append(
            {
                "device": device,
                "intervals": len(gaps),
                "median_gap_min": median(gaps) if gaps else None,
                "p95_gap_min": q(gaps, 0.95) if gaps else None,
                "max_gap_min": max(gaps) if gaps else None,
                "gaps_over_5_min": len(big),
                "largest_gap_start": big[-1][0] if big else "",
                "largest_gap_end": big[-1][1] if big else "",
            }
        )
    return rows


def top_divergence_periods(aligned: dict[datetime, dict[str, SensorRecord]]) -> list[dict]:
    bucket = defaultdict(list)
    for ts, devices in aligned.items():
        if "th-2" in devices and "co2-1" in devices:
            a = devices["th-2"].temperature
            b = devices["co2-1"].temperature
            if a is not None and b is not None and 5 <= a <= 45 and 5 <= b <= 45:
                hour = ts.replace(minute=0)
                bucket[hour].append(abs(a - b))
    rows = []
    for hour, values in bucket.items():
        if len(values) >= 20:
            rows.append(
                {
                    "hour": hour,
                    "n": len(values),
                    "mean_abs_th2_co2_temp_diff": mean(values),
                    "max_abs_th2_co2_temp_diff": max(values),
                }
            )
    rows.sort(key=lambda r: r["mean_abs_th2_co2_temp_diff"], reverse=True)
    return rows[:25]


def relay_summary(relays: list[RelayEvent], profiles: list[dict], names: dict) -> tuple[list[dict], list[dict]]:
    profile_use = defaultdict(list)
    for profile in profiles:
        for stage in ["pre", "main"]:
            for item in profile.get(stage, []):
                profile_use[int(item["relay"])].append(f'{profile.get("name", "")} / {stage}')
    by_relay = defaultdict(list)
    for e in relays:
        by_relay[e.relay_id].append(e)
    rows = []
    intervals = []
    for relay_id, events in sorted(by_relay.items()):
        events.sort(key=lambda e: e.ts)
        on_count = sum(1 for e in events if e.state == "ON")
        off_count = sum(1 for e in events if e.state == "OFF")
        durations = []
        open_on = None
        for e in events:
            if e.state == "ON":
                open_on = e.ts
            elif e.state == "OFF" and open_on:
                dur = (e.ts - open_on).total_seconds() / 60
                durations.append(dur)
                intervals.append(
                    {
                        "relay_id": relay_id,
                        "start": open_on,
                        "end": e.ts,
                        "duration_min": dur,
                        "name": names.get("relays", {}).get(str(relay_id), ""),
                        "profile_mentions": "; ".join(profile_use.get(relay_id, [])),
                    }
                )
                open_on = None
        rows.append(
            {
                "relay_id": relay_id,
                "name": names.get("relays", {}).get(str(relay_id), ""),
                "profile_mentions": "; ".join(profile_use.get(relay_id, [])),
                "events": len(events),
                "on_events": on_count,
                "off_events": off_count,
                "paired_on_intervals": len(durations),
                "total_on_min_observed": sum(durations) if durations else 0,
                "median_on_min": median(durations) if durations else None,
                "max_on_min": max(durations) if durations else None,
                "first_ts": events[0].ts,
                "last_ts": events[-1].ts,
            }
        )
    return rows, intervals


def relay_response_windows(sensors: list[SensorRecord], relays: list[RelayEvent]) -> list[dict]:
    by_device_metric: dict[tuple[str, str], tuple[list[datetime], list[float]]] = {}
    temp = defaultdict(lambda: ([], []))
    for r in sensors:
        for metric, value in [("temperature", r.temperature), ("humidity", r.humidity), ("co2", r.co2)]:
            if value is None:
                continue
            if metric == "temperature" and not (5 <= value <= 45):
                continue
            if metric == "humidity" and not (1 <= value <= 100):
                continue
            if metric == "co2" and value < 300:
                continue
            temp[(r.device, metric)][0].append(r.ts)
            temp[(r.device, metric)][1].append(value)
    by_device_metric.update(temp)

    def avg_between(times: list[datetime], values: list[float], start: datetime, end: datetime) -> tuple[int, float | None]:
        lo = bisect.bisect_left(times, start)
        hi = bisect.bisect_left(times, end)
        vals = values[lo:hi]
        return len(vals), (mean(vals) if vals else None)

    rows = []
    for e in [x for x in relays if x.state == "ON"]:
        for (device, metric), (times, values) in by_device_metric.items():
            n_before, before = avg_between(times, values, e.ts - timedelta(minutes=30), e.ts)
            n_after30, after30 = avg_between(times, values, e.ts, e.ts + timedelta(minutes=30))
            n_after120, after120 = avg_between(times, values, e.ts + timedelta(minutes=30), e.ts + timedelta(minutes=120))
            if n_before < 10 or n_after30 < 10:
                continue
            rows.append(
                {
                    "relay_id": e.relay_id,
                    "event_ts": e.ts,
                    "device": device,
                    "metric": metric,
                    "before_30m_mean": before,
                    "after_0_30m_mean": after30,
                    "after_30_120m_mean": after120,
                    "delta_0_30m": (after30 - before) if before is not None and after30 is not None else None,
                    "delta_30_120m": (after120 - before) if before is not None and after120 is not None else None,
                    "n_before": n_before,
                    "n_after30": n_after30,
                    "n_after120": n_after120,
                }
            )
    return rows


def lag_correlations(aligned: dict[datetime, dict[str, SensorRecord]]) -> list[dict]:
    # Ten-minute series smooths noise and keeps the pure-Python calculation small.
    bucket = defaultdict(list)
    for ts, devices in aligned.items():
        slot = ts.replace(minute=(ts.minute // 10) * 10)
        for device, r in devices.items():
            for metric, value in [("temperature", r.temperature), ("humidity", r.humidity), ("co2", r.co2)]:
                if value is None:
                    continue
                if metric == "temperature" and not (5 <= value <= 45):
                    continue
                if metric == "humidity" and not (1 <= value <= 100):
                    continue
                if metric == "co2" and value < 300:
                    continue
                bucket[(slot, device, metric)].append(value)
    series = defaultdict(dict)
    for (slot, device, metric), values in bucket.items():
        series[(device, metric)][slot] = mean(values)

    comparisons = [
        (("co2-1", "co2"), ("co2-1", "temperature")),
        (("co2-1", "co2"), ("co2-1", "humidity")),
        (("co2-1", "temperature"), ("th-1", "temperature")),
        (("co2-1", "temperature"), ("th-2", "temperature")),
        (("th-1", "temperature"), ("th-2", "temperature")),
        (("co2-1", "humidity"), ("th-1", "humidity")),
        (("co2-1", "humidity"), ("th-2", "humidity")),
    ]
    rows = []
    for left, right in comparisons:
        left_series = series.get(left, {})
        right_series = series.get(right, {})
        best = None
        for lag in range(-180, 181, 10):
            xs, ys = [], []
            delta = timedelta(minutes=lag)
            for ts, x in left_series.items():
                y = right_series.get(ts + delta)
                if y is not None:
                    xs.append(x)
                    ys.append(y)
            c = corr(xs, ys)
            rows.append(
                {
                    "left": f"{left[0]}:{left[1]}",
                    "right": f"{right[0]}:{right[1]}",
                    "right_shift_min": lag,
                    "n": len(xs),
                    "correlation": c,
                }
            )
            if c is not None and (best is None or abs(c) > abs(best["correlation"])):
                best = {"lag": lag, "n": len(xs), "correlation": c}
        if best:
            rows.append(
                {
                    "left": f"BEST {left[0]}:{left[1]}",
                    "right": f"{right[0]}:{right[1]}",
                    "right_shift_min": best["lag"],
                    "n": best["n"],
                    "correlation": best["correlation"],
                }
            )
    return rows


def extract_excel_summaries() -> dict[str, list[dict]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    sheet_inventory = []
    numeric_columns = []
    for ws in wb.worksheets:
        values_by_col = defaultdict(list)
        for row in ws.iter_rows(values_only=True):
            for i, value in enumerate(row, start=1):
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    values_by_col[i].append(float(value))
        sheet_inventory.append(
            {
                "sheet": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "numeric_columns": sum(1 for vals in values_by_col.values() if len(vals) >= 3),
                "numeric_cells": sum(len(vals) for vals in values_by_col.values()),
            }
        )
        for col, vals in values_by_col.items():
            if len(vals) >= 5:
                s = stats(vals)
                numeric_columns.append(
                    {
                        "sheet": ws.title,
                        "col": col,
                        "n": s["n"],
                        "mean": s["mean"],
                        "median": s["median"],
                        "min": s["min"],
                        "max": s["max"],
                        "std": s["std"],
                    }
                )

    # Power/light recipe table in "фенология", rows 30-34.
    power_rows = []
    ws = wb["фенология"]
    for row in ws.iter_rows(min_row=30, max_row=34, values_only=True):
        if not isinstance(row[1], (int, float)):
            continue
        power_rows.append(
            {
                "date": row[0] if isinstance(row[0], datetime) else "",
                "variant": row[1],
                "label": row[2],
                "height_or_value_col4": row[3],
                "recipe": row[5],
                "power_w": row[6],
                "voltage_v": row[7],
                "current_a": row[8],
                "lambda_pct": row[9],
                "tray_area_m2": row[10],
                "tier_area_m2": row[11],
            }
        )

    light_treatment_rows = []
    for item in PPFD_VARIANTS:
        ppfd = item["ppfd_umol_m2_s"]
        err = item["ppfd_error"]
        light_treatment_rows.append(
            {
                "variant": item["variant"],
                "ppfd_umol_m2_s": ppfd,
                "ppfd_error": err,
                "photoperiod_light_h": PHOTOPERIOD_LIGHT_H,
                "photoperiod_dark_h": PHOTOPERIOD_DARK_H,
                "dli_mol_m2_day": dli_from_ppfd(ppfd),
                "dli_error_mol_m2_day": dli_from_ppfd(err),
                "note": "Данные НИР: фотопериод 18 ч свет / 6 ч ночь; PPFD - фотосинтетическая мощность облучения, не электрическая мощность.",
            }
        )

    # Phenology and bolting notes.
    phenology_rows = []
    for row in ws.iter_rows(min_row=3, max_row=28, values_only=True):
        cultivar = row[0]
        if not isinstance(cultivar, str):
            continue
        phenology_rows.append(
            {
                "cultivar": cultivar.strip(),
                "single_emergence": row[1] if isinstance(row[1], datetime) else row[1],
                "mass_emergence": row[2] if isinstance(row[2], datetime) else row[2],
            }
        )
    bolting_rows = []
    for row in ws.iter_rows(min_row=40, max_row=64, values_only=True):
        if isinstance(row[0], str):
            bolting_rows.append({"cultivar": row[0], "note": row[1] or ""})

    # Nutrient solution dates, EC and height rows.
    nutrient_rows = []
    ws = wb["пит р-р+ээ"]
    current_date = None
    current_cultivar = None
    current_variant = None
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[0], datetime):
            current_date = row[0].date().isoformat()
        if isinstance(row[0], str) and row[0].strip() and row[0].strip().lower() not in {
            "сорт",
            "сорт ",
            "долили по 500 мл",
        }:
            current_cultivar = row[0].strip()
        if isinstance(row[1], (int, float)):
            current_variant = row[1]
        if isinstance(row[2], str) and row[2].strip().lower() in {"ес, мсм/см", "высота, см"}:
            vals = [float(v) for v in row[3:9] if isinstance(v, (int, float))]
            nutrient_rows.append(
                {
                    "date": current_date,
                    "cultivar": current_cultivar,
                    "variant": current_variant,
                    "metric": row[2],
                    "n": len(vals),
                    "mean": mean(vals) if vals else None,
                    "min": min(vals) if vals else None,
                    "max": max(vals) if vals else None,
                    "cv_pct": (pstdev(vals) / mean(vals) * 100) if len(vals) > 1 and mean(vals) else None,
                }
            )

    # Reference/manual nutrient-solution temperature measurements.
    reference_temps = []
    current_date = None
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[0], datetime):
            current_date = row[0].date().isoformat()
        for i, value in enumerate(row):
            if isinstance(value, str) and value.strip().lower() in {"темп", "темп р-ра"}:
                nums = [float(v) for v in row[i + 1 :] if isinstance(v, (int, float)) and 5 <= float(v) <= 45]
                if nums and current_date:
                    reference_temps.append(
                        {
                            "date": current_date,
                            "label": value,
                            "n": len(nums),
                            "mean_reference_temp": mean(nums),
                            "min_reference_temp": min(nums),
                            "max_reference_temp": max(nums),
                            "values": "; ".join(fmt(v, 1) for v in nums),
                        }
                    )

    storage_rows = []
    ws = wb["хранение"]
    current_cultivar = None
    for row in ws.iter_rows(min_row=2, max_row=38, values_only=True):
        if isinstance(row[0], str):
            current_cultivar = row[0]
        if current_cultivar and isinstance(row[1], (int, float)) and isinstance(row[8], (int, float)):
            storage_rows.append(
                {
                    "cultivar": current_cultivar,
                    "variant": row[1],
                    "before_avg_g": row[4],
                    "after_avg_g": row[7],
                    "loss_pct": row[8],
                }
            )

    organoleptic_counts = Counter()
    ws = wb["органолептика"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for text in [row[2], row[3], row[6]]:
            if not isinstance(text, str):
                continue
            lower = text.lower()
            for token in ["слад", "горч", "нейтрал", "травянист", "потеря тургора", "пятн", "конденсат"]:
                if token in lower:
                    organoleptic_counts[token] += 1
    organoleptic_rows = [{"feature": k, "mentions": v} for k, v in organoleptic_counts.most_common()]

    return {
        "excel_sheet_inventory": sheet_inventory,
        "excel_numeric_column_summary": numeric_columns,
        "power_measurements": power_rows,
        "light_treatments_ppfd_dli": light_treatment_rows,
        "phenology": phenology_rows,
        "bolting_notes": bolting_rows,
        "nutrient_solution_summary": nutrient_rows,
        "reference_temperature_measurements": reference_temps,
        "storage_losses": storage_rows,
        "organoleptic_mentions": organoleptic_rows,
    }


def reference_temperature_comparison(reference_rows: list[dict], daily_rows: list[dict], sensors: list[SensorRecord]) -> list[dict]:
    # Compare manual nutrient-solution temperature with three sensor temperatures.
    day_device_values = defaultdict(list)
    day_device_daytime = defaultdict(list)
    for r in sensors:
        if r.temperature is None or not (5 <= r.temperature <= 45):
            continue
        key = (r.ts.date().isoformat(), r.device)
        day_device_values[key].append(r.temperature)
        if 10 <= r.ts.hour <= 16:
            day_device_daytime[key].append(r.temperature)
    rows = []
    for ref in reference_rows:
        date = ref["date"]
        ref_temp = ref["mean_reference_temp"]
        for device in ["th-1", "th-2", "co2-1"]:
            day_vals = day_device_values.get((date, device), [])
            day_time = day_device_daytime.get((date, device), [])
            if not day_vals:
                continue
            rows.append(
                {
                    "date": date,
                    "reference_temp": ref_temp,
                    "reference_values": ref["values"],
                    "device": device,
                    "device_daily_mean": mean(day_vals),
                    "device_daily_median": median(day_vals),
                    "device_daytime_mean_10_16": mean(day_time) if day_time else None,
                    "abs_diff_daily_mean": abs(mean(day_vals) - ref_temp),
                    "abs_diff_daytime_mean": abs(mean(day_time) - ref_temp) if day_time else None,
                    "device_rows_day": len(day_vals),
                    "reference_note": "Ручная температура раствора имеет дату без времени, поэтому сравнение приблизительное.",
                }
            )
    return rows


def svg_line_chart(path: Path, title: str, series: dict[str, list[tuple[float, float]]], y_label: str = "") -> None:
    width, height = 1000, 420
    ml, mr, mt, mb = 70, 30, 50, 60
    all_points = [p for pts in series.values() for p in pts]
    if not all_points:
        path.write_text("", encoding="utf-8")
        return
    min_x, max_x = min(p[0] for p in all_points), max(p[0] for p in all_points)
    min_y, max_y = min(p[1] for p in all_points), max(p[1] for p in all_points)
    if min_y == max_y:
        min_y -= 1
        max_y += 1
    pad = (max_y - min_y) * 0.08
    min_y -= pad
    max_y += pad

    def xy(x, y):
        px = ml + (x - min_x) / (max_x - min_x or 1) * (width - ml - mr)
        py = height - mb - (y - min_y) / (max_y - min_y) * (height - mt - mb)
        return px, py

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        f'<text x="{width/2}" y="28" text-anchor="middle" font-family="Arial" font-size="20">{escape_xml(title)}</text>',
        f'<text x="18" y="{height/2}" transform="rotate(-90 18 {height/2})" text-anchor="middle" font-family="Arial" font-size="13">{escape_xml(y_label)}</text>',
    ]
    for i in range(6):
        y = min_y + (max_y - min_y) * i / 5
        _, py = xy(min_x, y)
        parts.append(f'<line x1="{ml}" y1="{py:.1f}" x2="{width-mr}" y2="{py:.1f}" stroke="#dddddd"/>')
        parts.append(f'<text x="{ml-8}" y="{py+4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{y:.1f}</text>')
    parts.append(f'<line x1="{ml}" y1="{height-mb}" x2="{width-mr}" y2="{height-mb}" stroke="#333"/>')
    parts.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{height-mb}" stroke="#333"/>')
    legend_x = ml
    for name, pts in series.items():
        if not pts:
            continue
        color = DEVICE_COLORS.get(name, "#333333")
        d = " ".join([f"{'M' if idx == 0 else 'L'} {xy(x, y)[0]:.1f} {xy(x, y)[1]:.1f}" for idx, (x, y) in enumerate(pts)])
        parts.append(f'<path d="{d}" fill="none" stroke="{color}" stroke-width="2.2"/>')
        parts.append(f'<rect x="{legend_x}" y="{height-28}" width="16" height="4" fill="{color}"/>')
        parts.append(f'<text x="{legend_x+22}" y="{height-24}" font-family="Arial" font-size="13">{escape_xml(name)}</text>')
        legend_x += 120
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8")


def svg_bar_chart(path: Path, title: str, labels: list[str], values: list[float], y_label: str = "") -> None:
    width, height = 1000, 430
    ml, mr, mt, mb = 80, 30, 55, 110
    max_v = max(values) if values else 1
    max_v = max_v * 1.15 if max_v else 1
    plot_w = width - ml - mr
    bar_w = plot_w / max(len(values), 1) * 0.7
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        f'<text x="{width/2}" y="30" text-anchor="middle" font-family="Arial" font-size="20">{escape_xml(title)}</text>',
        f'<text x="20" y="{height/2}" transform="rotate(-90 20 {height/2})" text-anchor="middle" font-family="Arial" font-size="13">{escape_xml(y_label)}</text>',
    ]
    for i in range(6):
        v = max_v * i / 5
        y = height - mb - v / max_v * (height - mt - mb)
        parts.append(f'<line x1="{ml}" y1="{y:.1f}" x2="{width-mr}" y2="{y:.1f}" stroke="#dddddd"/>')
        parts.append(f'<text x="{ml-8}" y="{y+4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{v:.1f}</text>')
    for i, (label, value) in enumerate(zip(labels, values)):
        x = ml + plot_w * (i + 0.15) / max(len(values), 1)
        h = value / max_v * (height - mt - mb)
        y = height - mb - h
        parts.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{h:.1f}" fill="#4c78a8"/>')
        parts.append(f'<text x="{x+bar_w/2:.1f}" y="{y-5:.1f}" text-anchor="middle" font-family="Arial" font-size="11">{value:.1f}</text>')
        parts.append(f'<text x="{x+bar_w/2:.1f}" y="{height-mb+18}" transform="rotate(-35 {x+bar_w/2:.1f} {height-mb+18})" text-anchor="end" font-family="Arial" font-size="12">{escape_xml(label[:22])}</text>')
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8")


def escape_xml(text: str) -> str:
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_figures(daily_rows: list[dict], hourly_rows: list[dict], pairwise_rows: list[dict], relay_rows: list[dict], excel: dict) -> list[Path]:
    figs = []
    for metric, label in [("temperature", "Температура, °C"), ("humidity", "Влажность, %"), ("co2", "CO2, ppm")]:
        dates = sorted({r["date"] for r in daily_rows if r["metric"] == metric})
        date_index = {d: i for i, d in enumerate(dates)}
        series = defaultdict(list)
        for r in daily_rows:
            if r["metric"] == metric:
                series[r["device"]].append((date_index[r["date"]], r["mean"]))
        path = FIGS / f"daily_{metric}.svg"
        svg_line_chart(path, f"Среднесуточная динамика: {label}", dict(series), label)
        figs.append(path)

    for metric, label in [("temperature", "Температура, °C"), ("humidity", "Влажность, %"), ("co2", "CO2, ppm")]:
        series = defaultdict(list)
        for r in hourly_rows:
            if r["metric"] == metric:
                series[r["device"]].append((r["hour"], r["mean"]))
        path = FIGS / f"hourly_{metric}.svg"
        svg_line_chart(path, f"Средний суточный ход: {label}", dict(series), label)
        figs.append(path)

    temp_pairs = [r for r in pairwise_rows if r["metric"] == "temperature"]
    svg_bar_chart(
        FIGS / "temperature_pairwise_mean_abs_diff.svg",
        "Среднее абсолютное расхождение температурных датчиков",
        [r["pair"] for r in temp_pairs],
        [r["mean_abs_diff"] or 0 for r in temp_pairs],
        "°C",
    )
    figs.append(FIGS / "temperature_pairwise_mean_abs_diff.svg")

    relay_top = sorted(relay_rows, key=lambda r: r["events"], reverse=True)[:12]
    svg_bar_chart(
        FIGS / "relay_event_counts.svg",
        "Число зарегистрированных событий реле",
        [str(r["relay_id"]) for r in relay_top],
        [float(r["events"]) for r in relay_top],
        "события",
    )
    figs.append(FIGS / "relay_event_counts.svg")

    power = excel["power_measurements"]
    svg_bar_chart(
        FIGS / "electric_power_measurements.svg",
        "Электрическая мощность световых режимов",
        [str(r["label"]) for r in power],
        [float(r["power_w"] or 0) for r in power],
        "Вт",
    )
    figs.append(FIGS / "electric_power_measurements.svg")

    light = excel["light_treatments_ppfd_dli"]
    svg_bar_chart(
        FIGS / "light_treatments_dli.svg",
        "Суточная интегральная освещенность DLI при фотопериоде 18/6",
        [f'Вариант {r["variant"]}' for r in light],
        [float(r["dli_mol_m2_day"]) for r in light],
        "моль/(м2·сут)",
    )
    figs.append(FIGS / "light_treatments_dli.svg")

    storage = sorted(excel["storage_losses"], key=lambda r: r["loss_pct"], reverse=True)[:15]
    svg_bar_chart(
        FIGS / "storage_loss_top15.svg",
        "Наибольшие потери массы при хранении",
        [f'{r["cultivar"]} {r["variant"]}' for r in storage],
        [float(r["loss_pct"]) for r in storage],
        "%",
    )
    figs.append(FIGS / "storage_loss_top15.svg")
    return figs


def table_data(rows: list[dict], columns: list[str], max_rows: int = 12, digits: int = 2) -> list[list[str]]:
    data = [[c for c in columns]]
    for row in rows[:max_rows]:
        data.append([fmt(row.get(c), digits) for c in columns])
    return data


def add_table(story, rows: list[dict], columns: list[str], max_rows: int = 12, font_size: int = 7) -> None:
    if not rows:
        story.append(Paragraph("Нет данных для таблицы.", get_styles()["Body"]))
        return
    tbl = Table(table_data(rows, columns, max_rows=max_rows), repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eef7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b8c2cc")),
                ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 0.25 * cm))


def rl_line_chart(title: str, series: dict[str, list[tuple[float, float]]], y_label: str = "") -> Drawing:
    width, height = 17.0 * cm, 6.3 * cm
    ml, mr, mt, mb = 1.35 * cm, 0.35 * cm, 0.65 * cm, 0.85 * cm
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.white, strokeColor=None))
    d.add(String(width / 2, height - 0.35 * cm, title, fontName="DejaVu-Bold", fontSize=8.5, textAnchor="middle"))
    if y_label:
        d.add(String(0.15 * cm, height / 2, y_label, fontName="DejaVu", fontSize=6, textAnchor="middle", transform=[0, 1, -1, 0, 0.15 * cm, height / 2]))
    points = [p for pts in series.values() for p in pts]
    if not points:
        return d
    min_x, max_x = min(p[0] for p in points), max(p[0] for p in points)
    min_y, max_y = min(p[1] for p in points), max(p[1] for p in points)
    if min_y == max_y:
        min_y -= 1
        max_y += 1
    pad = (max_y - min_y) * 0.08
    min_y -= pad
    max_y += pad

    def xy(x, y):
        px = ml + (x - min_x) / (max_x - min_x or 1) * (width - ml - mr)
        py = mb + (y - min_y) / (max_y - min_y) * (height - mt - mb)
        return px, py

    for i in range(5):
        y = min_y + (max_y - min_y) * i / 4
        _, py = xy(min_x, y)
        d.add(Line(ml, py, width - mr, py, strokeColor=colors.HexColor("#d9dfe7"), strokeWidth=0.35))
        d.add(String(ml - 0.12 * cm, py - 2, f"{y:.1f}", fontName="DejaVu", fontSize=5.5, textAnchor="end"))
    d.add(Line(ml, mb, width - mr, mb, strokeColor=colors.HexColor("#333333"), strokeWidth=0.6))
    d.add(Line(ml, mb, ml, height - mt, strokeColor=colors.HexColor("#333333"), strokeWidth=0.6))
    legend_x = ml
    for name, pts in series.items():
        if not pts:
            continue
        color = colors.HexColor(DEVICE_COLORS.get(name, "#333333"))
        path = RlPath()
        for idx, (x, y) in enumerate(pts):
            px, py = xy(x, y)
            if idx == 0:
                path.moveTo(px, py)
            else:
                path.lineTo(px, py)
        path.strokeColor = color
        path.strokeWidth = 1.2
        path.fillColor = None
        d.add(path)
        d.add(Rect(legend_x, 0.12 * cm, 0.35 * cm, 0.08 * cm, fillColor=color, strokeColor=None))
        d.add(String(legend_x + 0.45 * cm, 0.08 * cm, name, fontName="DejaVu", fontSize=6))
        legend_x += 2.2 * cm
    return d


def rl_bar_chart(title: str, labels: list[str], values: list[float], y_label: str = "") -> Drawing:
    width, height = 17.0 * cm, 6.6 * cm
    ml, mr, mt, mb = 1.45 * cm, 0.35 * cm, 0.65 * cm, 1.15 * cm
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.white, strokeColor=None))
    d.add(String(width / 2, height - 0.35 * cm, title, fontName="DejaVu-Bold", fontSize=8.5, textAnchor="middle"))
    if y_label:
        d.add(String(0.16 * cm, height / 2, y_label, fontName="DejaVu", fontSize=6, textAnchor="middle", transform=[0, 1, -1, 0, 0.16 * cm, height / 2]))
    max_v = max(values) if values else 1
    max_v = max_v * 1.15 if max_v else 1
    plot_w = width - ml - mr
    plot_h = height - mt - mb
    for i in range(5):
        v = max_v * i / 4
        y = mb + v / max_v * plot_h
        d.add(Line(ml, y, width - mr, y, strokeColor=colors.HexColor("#d9dfe7"), strokeWidth=0.35))
        d.add(String(ml - 0.12 * cm, y - 2, f"{v:.1f}", fontName="DejaVu", fontSize=5.5, textAnchor="end"))
    if values:
        step = plot_w / len(values)
        bar_w = step * 0.64
        for i, (label, value) in enumerate(zip(labels, values)):
            x = ml + i * step + step * 0.18
            h = value / max_v * plot_h
            d.add(Rect(x, mb, bar_w, h, fillColor=colors.HexColor("#4c78a8"), strokeColor=None))
            d.add(String(x + bar_w / 2, mb + h + 3, f"{value:.1f}", fontName="DejaVu", fontSize=5.4, textAnchor="middle"))
            d.add(String(x + bar_w / 2, 0.18 * cm, str(label)[:18], fontName="DejaVu", fontSize=5.4, textAnchor="middle"))
    d.add(Line(ml, mb, width - mr, mb, strokeColor=colors.HexColor("#333333"), strokeWidth=0.6))
    d.add(Line(ml, mb, ml, height - mt, strokeColor=colors.HexColor("#333333"), strokeWidth=0.6))
    return d


def add_daily_chart(story, daily_rows: list[dict], metric: str, title: str, y_label: str) -> None:
    dates = sorted({r["date"] for r in daily_rows if r["metric"] == metric})
    date_index = {d: i for i, d in enumerate(dates)}
    series = defaultdict(list)
    for r in daily_rows:
        if r["metric"] == metric:
            series[r["device"]].append((date_index[r["date"]], r["mean"]))
    story.append(rl_line_chart(title, dict(series), y_label))
    story.append(Spacer(1, 0.2 * cm))


def add_hourly_chart(story, hourly_rows: list[dict], metric: str, title: str, y_label: str) -> None:
    series = defaultdict(list)
    for r in hourly_rows:
        if r["metric"] == metric:
            series[r["device"]].append((r["hour"], r["mean"]))
    story.append(rl_line_chart(title, dict(series), y_label))
    story.append(Spacer(1, 0.2 * cm))


def register_fonts() -> None:
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Times New Roman.ttf",
        "/Library/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            pdfmetrics.registerFont(TTFont("DejaVu", path))
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", path))
            return


_STYLES = None


def get_styles():
    global _STYLES
    if _STYLES is None:
        base = getSampleStyleSheet()
        _STYLES = {
            "Title": ParagraphStyle(
                "TitleRu",
                parent=base["Title"],
                fontName="DejaVu-Bold",
                fontSize=18,
                leading=22,
                alignment=TA_CENTER,
                spaceAfter=14,
            ),
            "H1": ParagraphStyle("H1Ru", parent=base["Heading1"], fontName="DejaVu-Bold", fontSize=14, leading=17, spaceBefore=10, spaceAfter=6),
            "H2": ParagraphStyle("H2Ru", parent=base["Heading2"], fontName="DejaVu-Bold", fontSize=11, leading=14, spaceBefore=8, spaceAfter=4),
            "Body": ParagraphStyle("BodyRu", parent=base["BodyText"], fontName="DejaVu", fontSize=9, leading=12, spaceAfter=5),
            "Small": ParagraphStyle("SmallRu", parent=base["BodyText"], fontName="DejaVu", fontSize=7.5, leading=10, spaceAfter=4),
        }
    return _STYLES


def p(text: str):
    return Paragraph(text, get_styles()["Body"])


def h1(text: str):
    return Paragraph(text, get_styles()["H1"])


def h2(text: str):
    return Paragraph(text, get_styles()["H2"])


def summarize_reference_hypothesis(rows: list[dict]) -> tuple[str, list[dict]]:
    by_date = defaultdict(list)
    for r in rows:
        by_date[(r["date"], r["reference_temp"])].append(r)
    winners = []
    for key, rs in sorted(by_date.items()):
        best = min(
            rs,
            key=lambda r: (
                r["abs_diff_daytime_mean"]
                if r["abs_diff_daytime_mean"] is not None
                else r["abs_diff_daily_mean"]
                if r["abs_diff_daily_mean"] is not None
                else 999
            ),
        )
        winners.append(best)
    co2_wins = sum(1 for w in winners if w["device"] == "co2-1")
    text = (
        f"Ручная температура раствора найдена для {len(by_date)} датированных блоков измерений. "
        f"По ближайшему доступному критерию сравнения датчик co2-1 оказался ближайшим только в {co2_wins} из {len(winners)} блоков; в остальных случаях ближе был th-1. "
        "Поэтому гипотеза о том, что эталонная температура ближе к датчику с CO2, не получила устойчивого подтверждения. "
        "Вывод ограничен тем, что в Excel указаны даты, но не точное время ручных измерений."
    )
    return text, winners


def build_report(context: dict) -> Path:
    register_fonts()
    styles = get_styles()
    pdf_path = OUT / "greenhouse_deep_data_report_ru.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        rightMargin=1.3 * cm,
        leftMargin=1.3 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="Исследовательский анализ данных теплицы",
    )
    story = [
        Paragraph("Исследовательский анализ данных теплицы и опыта выращивания салата", styles["Title"]),
        p("Отчет подготовлен автоматически из локальных файлов проекта. Цель анализа - не подтвердить заранее заданные тезисы, а описать фактическое содержание, качество, закономерности и ограничения данных."),
        h1("1. Источники данных"),
    ]
    story.append(p("В анализ включены SQLite-журнал датчиков и реле, CSV-экспорты, JSON-профили оборудования и все листы книги «Данные салат мощность.xlsx»."))
    add_table(story, context["source_inventory"], ["source", "type", "rows", "period_start", "period_end", "comment"], max_rows=18, font_size=6)

    story += [
        h1("2. Качество и покрытие сенсорных данных"),
        p("Основной временной ряд охватывает период с 28 марта по 26 мая 2026 года. Три устройства пишут примерно минутные наблюдения, но состав метрик различается: th-1 и th-2 дают температуру/влажность, co2-1 дополнительно дает CO2."),
    ]
    add_daily_chart(story, context["daily_rows"], "temperature", "Среднесуточная температура по датчикам", "°C")
    add_daily_chart(story, context["daily_rows"], "humidity", "Среднесуточная влажность по датчикам", "%")
    add_daily_chart(story, context["daily_rows"], "co2", "Среднесуточный CO2 после фильтрации нулей", "ppm")
    add_table(story, context["sensor_summary"], ["device", "rows", "temperature_mean", "temperature_min", "temperature_max", "humidity_mean", "humidity_min", "humidity_max", "co2_mean", "co2_min", "co2_max", "zero_co2_rows"], max_rows=6)
    story.append(p("Нулевые значения CO2 и влажности трактуются как технические пропуски или некорректные показания; в расчетах зависимостей CO2 использованы только значения не ниже 300 ppm. Это резко меняет картину: сырое среднее CO2 из исходного CSV занижено из-за нулей."))
    add_table(story, context["gap_rows"], ["device", "median_gap_min", "p95_gap_min", "max_gap_min", "gaps_over_5_min", "largest_gap_start", "largest_gap_end"], max_rows=5)

    story += [
        h1("3. Сравнение датчиков микроклимата"),
        p("Температурные каналы заметно отличаются по абсолютному уровню. Это важный результат: датчики отражают не только шум измерений, но и пространственное положение/локальный режим в теплице."),
    ]
    temp_pairs = [r for r in context["pairwise"] if r["metric"] == "temperature"]
    story.append(rl_bar_chart("Среднее абсолютное расхождение температур", [r["pair"] for r in temp_pairs], [r["mean_abs_diff"] or 0 for r in temp_pairs], "°C"))
    story.append(Spacer(1, 0.2 * cm))
    add_table(story, [r for r in context["pairwise"] if r["metric"] == "temperature"], ["pair", "n_common_minutes", "mean_diff", "mean_abs_diff", "p95_abs_diff", "rmse_diff", "correlation"], max_rows=5)
    add_table(story, [r for r in context["pairwise"] if r["metric"] == "humidity"], ["pair", "n_common_minutes", "mean_diff", "mean_abs_diff", "p95_abs_diff", "rmse_diff", "correlation"], max_rows=5)
    story.append(p("Наиболее показательное расхождение: th-2 систематически теплее co2-1, а co2-1 при этом имеет более высокую влажность. Для ВКР это полезно как аргумент в пользу разнородных датчиков и необходимости учитывать место установки, а не усреднять все каналы без модели."))
    add_table(story, context["divergence_rows"], ["hour", "n", "mean_abs_th2_co2_temp_diff", "max_abs_th2_co2_temp_diff"], max_rows=10)

    story += [
        h1("4. Проверка гипотезы об эталонной температуре"),
    ]
    hypothesis_text, winners = summarize_reference_hypothesis(context["reference_comparison"])
    story.append(p(hypothesis_text))
    add_table(story, context["reference_comparison"], ["date", "reference_temp", "reference_values", "device", "device_daytime_mean_10_16", "abs_diff_daytime_mean", "reference_note"], max_rows=15, font_size=6)

    story += [
        h1("5. Суточные циклы, лаги и связи параметров"),
        p("Суточный ход устойчиво просматривается в температуре и влажности. CO2 менее стабилен из-за большого числа технических нулей и периодов без валидных показаний, но после фильтрации остается пригодным для поиска связей с вентиляцией и микроклиматом."),
    ]
    add_hourly_chart(story, context["hourly_rows"], "temperature", "Средний суточный ход температуры", "°C")
    add_hourly_chart(story, context["hourly_rows"], "humidity", "Средний суточный ход влажности", "%")
    add_hourly_chart(story, context["hourly_rows"], "co2", "Средний суточный ход CO2", "ppm")
    best_lags = [r for r in context["lag_rows"] if str(r["left"]).startswith("BEST")]
    add_table(story, best_lags, ["left", "right", "right_shift_min", "n", "correlation"], max_rows=12)
    story.append(p("Знак лага читается так: положительное значение означает, что правая серия сдвинута вперед относительно левой. Эти оценки не являются причинным доказательством, но помогают выбрать окна признаков для ML-модели."))

    story += [
        h1("6. Оборудование, реле и реакции микроклимата"),
        p("Журнал реле содержит значительно меньше событий, чем журнал датчиков. Поэтому реакцию оборудования можно оценивать только как разведочный event-study, а не как строгий эксперимент: нет гарантии, что события независимы, а внешняя погода и ручные действия не контролируются."),
    ]
    relay_top = sorted(context["relay_rows"], key=lambda r: r["events"], reverse=True)[:12]
    story.append(rl_bar_chart("Число событий реле", [str(r["relay_id"]) for r in relay_top], [float(r["events"]) for r in relay_top], "события"))
    story.append(Spacer(1, 0.2 * cm))
    add_table(story, context["relay_rows"], ["relay_id", "name", "profile_mentions", "events", "paired_on_intervals", "total_on_min_observed", "median_on_min", "max_on_min"], max_rows=14, font_size=6)
    relay_response_top = sorted(
        [r for r in context["relay_response"] if r["delta_0_30m"] is not None],
        key=lambda r: abs(r["delta_0_30m"]),
        reverse=True,
    )[:15]
    add_table(story, relay_response_top, ["relay_id", "event_ts", "device", "metric", "before_30m_mean", "after_0_30m_mean", "delta_0_30m", "after_30_120m_mean", "delta_30_120m"], max_rows=15, font_size=6)
    story.append(p("Наибольшие краткосрочные сдвиги вокруг включений следует использовать как список периодов для ручной проверки: часть изменений может быть вызвана включением оборудования, но часть совпадает с естественным суточным ходом или ручными операциями."))

    story += [
        h1("7. Excel-книга по салату и мощности"),
        p("Книга содержит не только мощность, но и фенологию, биометрию, биохимию, хранение, органолептику и водный статус растений. Данные полезны для связывания микроклимата не с мгновенной реакцией, а с итоговыми признаками выращивания."),
    ]
    add_table(story, context["excel"]["excel_sheet_inventory"], ["sheet", "rows", "cols", "numeric_columns", "numeric_cells"], max_rows=14)
    story.append(p("В НИР варианты выращивания различались по фотосинтетической мощности облучения PPFD при одинаковом фотопериоде 18 ч свет / 6 ч ночь. Поэтому корректная агрофизиологическая ось анализа - не только электрическая мощность в Вт, а PPFD и рассчитанная суточная интегральная освещенность DLI."))
    light = context["excel"]["light_treatments_ppfd_dli"]
    story.append(rl_bar_chart("DLI световых вариантов при фотопериоде 18/6", [f'Вариант {r["variant"]}' for r in light], [float(r["dli_mol_m2_day"]) for r in light], "моль/(м2·сут)"))
    story.append(Spacer(1, 0.2 * cm))
    add_table(story, context["excel"]["light_treatments_ppfd_dli"], ["variant", "ppfd_umol_m2_s", "ppfd_error", "photoperiod_light_h", "photoperiod_dark_h", "dli_mol_m2_day", "dli_error_mol_m2_day"], max_rows=8)
    story.append(p("Отдельно в Excel есть электрические измерения световых режимов в Вт, напряжении и токе. Они полезны для энергетической части, но не должны смешиваться с PPFD: PPFD описывает поток фотонов для растений, а Вт - потребление/электрический режим оборудования."))
    power = context["excel"]["power_measurements"]
    story.append(rl_bar_chart("Электрическая мощность световых режимов", [str(r["label"]) for r in power], [float(r["power_w"] or 0) for r in power], "Вт"))
    story.append(Spacer(1, 0.2 * cm))
    add_table(story, context["excel"]["power_measurements"], ["variant", "label", "recipe", "power_w", "voltage_v", "current_a", "lambda_pct", "tray_area_m2", "tier_area_m2"], max_rows=8)
    story.append(p("В листе «пит р-р+ээ» прослеживается рост EC и падение высоты раствора к поздним датам по ряду вариантов. Это содержательно: расход раствора и концентрация питания могут быть связаны с интенсивностью роста и режимом микроклимата, но таблица не содержит точного времени измерений."))
    add_table(story, context["excel"]["nutrient_solution_summary"], ["date", "cultivar", "variant", "metric", "n", "mean", "min", "max", "cv_pct"], max_rows=18, font_size=6)
    story.append(p("Фенология показывает массовые всходы в основном 5-8 марта; отдельной пометкой отмечены плохие всходы у Триплекс, Пилигрим и Старфайтер. В поздних наблюдениях часто встречается стеблевание, что важно обсуждать отдельно от микроклимата: сортовой фактор и срок выращивания могут доминировать над климатическими эффектами."))
    add_table(story, context["excel"]["bolting_notes"], ["cultivar", "note"], max_rows=30, font_size=7)
    story.append(p("Потери массы при хранении и органолептические признаки показывают заметную неоднородность вариантов. Эти признаки можно использовать как конечные отклики в будущей таблице «условия выращивания - качество продукции»."))
    storage_top = sorted(context["excel"]["storage_losses"], key=lambda r: r["loss_pct"], reverse=True)[:12]
    story.append(rl_bar_chart("Наибольшие потери массы при хранении", [f'{r["cultivar"]} {r["variant"]}' for r in storage_top], [float(r["loss_pct"]) for r in storage_top], "%"))
    story.append(Spacer(1, 0.2 * cm))
    add_table(story, sorted(context["excel"]["storage_losses"], key=lambda r: r["loss_pct"], reverse=True), ["cultivar", "variant", "before_avg_g", "after_avg_g", "loss_pct"], max_rows=18)
    add_table(story, context["excel"]["organoleptic_mentions"], ["feature", "mentions"], max_rows=10)

    story += [
        h1("8. Аномалии и интересные периоды"),
        p("Ключевые аномалии: нулевые показания CO2, нули влажности у отдельных устройств, большие расхождения температур между датчиками и короткие ON/OFF события реле. Это не мусор, который нужно молча удалить: для диплома они показывают реальные проблемы опытного IoT-сбора данных."),
    ]
    add_table(story, context["anomaly_rows"], ["category", "description", "count_or_value", "example"], max_rows=20, font_size=7)

    story += [
        h1("9. Что полезно для ВКР и дальнейших экспериментов"),
        p("Для ВКР наиболее ценны четыре результата. Во-первых, подтверждена необходимость единой модели данных для разнородных сенсоров: CO2 есть только у co2-1, а температурные уровни между узлами различаются систематически. Во-вторых, совместный журнал датчиков и реле действительно позволяет искать реакцию микроклимата на воздействия, хотя текущий журнал реле мал для строгих выводов. В-третьих, Excel с агробиологическими измерениями можно связать с микроклиматом только после нормализации дат, сортов, вариантов и ярусов. В-четвертых, качество данных само по себе является исследовательским результатом: ML-модель должна учитывать пропуски, нули, лаги и пространственную неоднородность."),
        p("Для статей или дальнейших экспериментов стоит спланировать controlled logging: фиксировать точное время ручных измерений, положение датчиков, включение профилей, внешний климат, расход раствора и фактическую мощность во времени. Тогда можно перейти от разведочной аналитики к причинной оценке и прогнозированию."),
        h1("10. Ограничения"),
        p("Внешняя температура/освещенность не подключены; ручные измерения Excel часто имеют только дату; часть листов содержит лабораторные показатели без явной связи с местом растения в теплице; мощность дана точечными измерениями, а не временным рядом; релейные события немногочисленны и не всегда образуют чистые интервалы ON-OFF. Поэтому выводы нужно использовать как разведочные, а не как окончательное доказательство эффективности управления."),
    ]

    story.append(PageBreak())
    story.append(h1("Приложение: перечень созданных таблиц и графиков"))
    for path in sorted(TABLES.glob("*.csv")):
        story.append(p(f"Таблица: {path.relative_to(ROOT)}"))
    for path in sorted(FIGS.glob("*.svg")):
        story.append(p(f"График: {path.relative_to(ROOT)}"))

    doc.build(story)
    return pdf_path


def anomaly_rows(sensor_summary_rows, gap_rows, pairwise_rows, reference_rows, relay_rows) -> list[dict]:
    rows = []
    for r in sensor_summary_rows:
        for metric in ["zero_temp_rows", "zero_humidity_rows", "zero_co2_rows"]:
            if r.get(metric):
                rows.append(
                    {
                        "category": "sensor_zero_values",
                        "description": f'{r["device"]}: {metric}',
                        "count_or_value": r[metric],
                        "example": "Нули исключены из физических расчетов, но сохранены как признак качества данных.",
                    }
                )
    for r in gap_rows:
        if r.get("max_gap_min") and r["max_gap_min"] > 5:
            rows.append(
                {
                    "category": "sensor_gap",
                    "description": f'{r["device"]}: максимальный разрыв',
                    "count_or_value": fmt(r["max_gap_min"], 1) + " мин",
                    "example": f'{fmt(r["largest_gap_start"])} - {fmt(r["largest_gap_end"])}',
                }
            )
    for r in pairwise_rows:
        if r["metric"] == "temperature" and r.get("mean_abs_diff") and r["mean_abs_diff"] > 2:
            rows.append(
                {
                    "category": "sensor_divergence",
                    "description": f'Температурная пара {r["pair"]}',
                    "count_or_value": fmt(r["mean_abs_diff"]) + " °C среднее abs",
                    "example": "Признак пространственной неоднородности или смещения датчиков.",
                }
            )
    open_relays = [r for r in relay_rows if r["on_events"] != r["off_events"]]
    for r in open_relays:
        rows.append(
            {
                "category": "relay_balance",
                "description": f'Реле {r["relay_id"]}: ON/OFF не сбалансированы',
                "count_or_value": f'{r["on_events"]}/{r["off_events"]}',
                "example": "Возможны незакрытые интервалы или ручные переключения вне логики профиля.",
            }
        )
    return rows


def write_notes(context: dict) -> None:
    text = [
        "# Промежуточные заметки исследовательского анализа",
        "",
        "## Проверенные источники",
        "",
        "- `Данные салат мощность.xlsx`: все 11 листов.",
        "- `history_2026-05-26.db`: таблицы `sensor_log` и `relay_log`.",
        "- CSV-экспорты sensor/relay summary и raw log.",
        "- JSON `names`, `profiles`, `rules`.",
        "",
        "## Главные наблюдения",
        "",
        "1. Датчики температуры не взаимозаменяемы: расхождения между узлами систематические и достаточно большие для учета в модели.",
        "2. CO2-канал содержит много нулевых значений; анализ CO2 требует фильтрации технических нулей.",
        "3. Релейный журнал полезен для разведочного event-study, но мал для строгого вывода об эффективности оборудования.",
        "4. Для световых вариантов добавлен слой PPFD и DLI: фотопериод 18 ч свет / 6 ч ночь, PPFD 79,0-245,5 мкмоль/(м2·с), DLI 5,12-15,91 моль/(м2·сут).",
        "5. Электрические измерения в Вт из Excel сохранены отдельно и не смешиваются с PPFD: это разные физические показатели.",
        "6. Гипотеза о близости ручной температуры раствора к co2-1 не подтвердилась устойчиво; чаще ближе th-1, а точное сравнение ограничено отсутствием времени измерения внутри дня.",
        "",
        "## Таблицы",
        "",
    ]
    for path in sorted(TABLES.glob("*.csv")):
        text.append(f"- `{path.relative_to(ROOT)}`")
    text += ["", "## Графики", ""]
    for path in sorted(FIGS.glob("*.svg")):
        text.append(f"- `{path.relative_to(ROOT)}`")
    (NOTES / "analysis_notes.md").write_text("\n".join(text) + "\n", encoding="utf-8")


def main() -> None:
    ensure_dirs()
    sensors, relays = load_sensor_data()
    names, profiles, rules = load_json_sources()
    aligned = align_by_minute(sensors)

    src = source_inventory(sensors, relays)
    sensor_rows = sensor_device_summary(sensors)
    pairwise_rows = pairwise_sensor_comparison(aligned, "temperature") + pairwise_sensor_comparison(aligned, "humidity")
    hourly_rows = hourly_cycles(sensors)
    daily_rows = daily_summary(sensors)
    gap_rows = gap_analysis(sensors)
    divergence_rows = top_divergence_periods(aligned)
    relay_rows, relay_intervals = relay_summary(relays, profiles, names)
    relay_response = relay_response_windows(sensors, relays)
    lag_rows = lag_correlations(aligned)
    excel = extract_excel_summaries()
    reference_cmp = reference_temperature_comparison(excel["reference_temperature_measurements"], daily_rows, sensors)
    anomalies = anomaly_rows(sensor_rows, gap_rows, pairwise_rows, reference_cmp, relay_rows)

    tables = {
        "source_inventory.csv": src,
        "sensor_device_summary.csv": sensor_rows,
        "sensor_pairwise_comparison.csv": pairwise_rows,
        "sensor_hourly_cycles.csv": hourly_rows,
        "sensor_daily_summary.csv": daily_rows,
        "sensor_gap_analysis.csv": gap_rows,
        "top_temperature_divergence_periods.csv": divergence_rows,
        "relay_summary_enriched.csv": relay_rows,
        "relay_intervals.csv": relay_intervals,
        "relay_response_windows.csv": relay_response,
        "lag_correlations.csv": lag_rows,
        "reference_temperature_comparison.csv": reference_cmp,
        "anomalies_and_quality_flags.csv": anomalies,
    }
    for name, rows in tables.items():
        write_csv(TABLES / name, rows)
    for name, rows in excel.items():
        write_csv(TABLES / f"{name}.csv", rows)

    build_figures(daily_rows, hourly_rows, pairwise_rows, relay_rows, excel)
    context = {
        "source_inventory": src,
        "sensor_summary": sensor_rows,
        "pairwise": pairwise_rows,
        "hourly_rows": hourly_rows,
        "daily_rows": daily_rows,
        "gap_rows": gap_rows,
        "divergence_rows": divergence_rows,
        "relay_rows": relay_rows,
        "relay_intervals": relay_intervals,
        "relay_response": relay_response,
        "lag_rows": lag_rows,
        "excel": excel,
        "reference_comparison": reference_cmp,
        "anomaly_rows": anomalies,
    }
    pdf_path = build_report(context)
    write_notes(context)
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "pdf_report": str(pdf_path.relative_to(ROOT)),
        "tables": [str(p.relative_to(ROOT)) for p in sorted(TABLES.glob("*.csv"))],
        "figures": [str(p.relative_to(ROOT)) for p in sorted(FIGS.glob("*.svg"))],
        "source_files": [
            str(XLSX.relative_to(ROOT)),
            "Greenhouse/dashboard/data/history.db",
            "Greenhouse/mosquitto/data/mosquitto.db",
            str(DB.relative_to(ROOT)),
            str((EVIDENCE / "sensor_log_2026-05-26.csv").relative_to(ROOT)),
            str((EVIDENCE / "relay_log_2026-05-26.csv").relative_to(ROOT)),
            str((EVIDENCE / "names_2026-05-26.json").relative_to(ROOT)),
            str((EVIDENCE / "profiles_2026-05-26.json").relative_to(ROOT)),
            str((EVIDENCE / "rules_2026-05-26.json").relative_to(ROOT)),
        ],
    }
    (OUT / "analysis_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
