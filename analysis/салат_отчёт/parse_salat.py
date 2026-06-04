# -*- coding: utf-8 -*-
"""Парсер агрономического массива по салату (Данные салат мощность.xlsx).

Извлекает данные в опрятные (tidy) таблицы:
  factorial  — основной опыт: 6 сортов x 5 световых режимов (средние по показателям)
  reps       — повторности по биометрии (для ошибок средних)
  screening  — сортовой скрининг (25 сортов, разовый замер): SPAD, биометрия, СВ, вода
  light_meta — параметры световых режимов (PPFD, мощность, DLI)
"""
from __future__ import annotations
import json
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Данные салат мощность.xlsx"
OUT = Path(__file__).resolve().parent
WB = load_workbook(XLSX, data_only=True)

CORE = ["Триплекс", "Рафаэль", "Джипси", "Цезарь", "Каскад", "Риф"]

# ---- Световые режимы (вариант 1..5) -------------------------------------
# PPFD взят из столбца "номинал" листа фенология; мощность/U/I/λ — измеренные.
LIGHT = pd.DataFrame({
    "variant":  [1, 2, 3, 4, 5],
    "name":     ["контроль", "ярус 1", "ярус 2", "ярус 3", "ярус 4"],
    "recipe":   ["контроль", "4000K-33%", "4000K-60%", "4000K-75%", "4000K-100%"],
    "ppfd":     [79.0, 95.0, 150.8, 181.3, 245.5],   # µmol·m^-2·s^-1 (номинал)
    "power_w":  [116.3, 97.0, 221.0, 321.5, 478.0],  # Вт (измеренная)
    "voltage":  [237.2, 236.3, 235.7, 235.3, 235.0],
    "current":  [0.514, 0.540, 1.020, 1.359, 2.053],
    "lambda":   [96.0, 77.0, 92.0, 98.0, 100.0],
})
LIGHT["dli"] = LIGHT["ppfd"] * 18 * 3600 / 1e6        # mol·m^-2·d^-1, фотопериод 18 ч
LIGHT["ppfd_per_w"] = LIGHT["ppfd"] / LIGHT["power_w"]


def _num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return np.nan


# ---- Хлорофиллметр (SPAD) -----------------------------------------------
def parse_spad():
    ws = WB["хлорофиллметр"]
    # левый блок: Сорт(1),Вариант(2),показания 3..8 ; ffill сорта
    rows = []
    cur = None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            vals = [_num(ws.cell(r, c).value) for c in range(3, 9)]
            vals = [v for v in vals if not np.isnan(v)]
            if vals:
                rows.append({"cultivar": cur, "variant": int(var),
                             "spad_mean": np.mean(vals), "spad_n": len(vals)})
    fac = pd.DataFrame(rows)
    # правый блок: скрининг — Сорт(10), показания 11..18
    scr = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 10).value
        if isinstance(name, str) and name.strip() and name != "Сорт":
            vals = [_num(ws.cell(r, c).value) for c in range(11, 19)]
            vals = [v for v in vals if not np.isnan(v)]
            if vals:
                scr.append({"cultivar": name.strip(), "spad_mean": np.mean(vals),
                            "spad_n": len(vals)})
    return fac, pd.DataFrame(scr)


# ---- Биометрия (8 блоков по 9 столбцов) ---------------------------------
BIO_BLOCKS = [
    (1,  "height_cm"),       # высота розетки / длина листа
    (10, "diam_cm"),         # диаметр розетки / ширина листа
    (19, "root_len_cm"),     # длина корней
    (28, "leaves_n"),        # кол-во листьев
    (37, "leaf_mass_mkt_g"), # масса листьев товарная
    (46, "leaf_mass_non_g"), # масса листьев нетоварная
    (55, "stem_mass_g"),     # масса стебля
    (64, "root_mass_g"),     # масса корня
]


def _bio_block(ws, c0, metric, cultivars, want_variant=True):
    """Возвращает long df: cultivar,variant,metric,value(reps mean)."""
    rows = []
    cur = None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, c0).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, c0 + 1).value
        if cur in cultivars and isinstance(var, (int, float)):
            reps = [_num(ws.cell(r, c).value) for c in range(c0 + 2, c0 + 8)]
            reps = [v for v in reps if not np.isnan(v)]
            if reps:
                rows.append({"cultivar": cur, "variant": int(var), metric: np.mean(reps),
                             metric + "_sd": (np.std(reps, ddof=1) if len(reps) > 1 else np.nan),
                             metric + "_n": len(reps)})
    return pd.DataFrame(rows)


def parse_biometry():
    ws = WB["биометрия"]
    # факторный опыт: только CORE сорта (первые ~31 строка), вариант 1..5
    dfs = []
    for c0, metric in BIO_BLOCKS:
        df = _bio_block(ws, c0, metric, CORE)
        df = df[df["variant"].between(1, 5)]
        dfs.append(df)
    fac = dfs[0]
    for df in dfs[1:]:
        fac = fac.merge(df, on=["cultivar", "variant"], how="outer")
    return fac


# ---- Биохимия -----------------------------------------------------------
def parse_biochem():
    ws = WB["биохимия"]
    rows = []
    cur = None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            rows.append({
                "cultivar": cur, "variant": int(var),
                # свежее: Ca(8) Cb(9) сумм(13) Xk-кар(14)
                "chl_a": _num(ws.cell(r, 8).value),
                "chl_b": _num(ws.cell(r, 9).value),
                "chl_sum": _num(ws.cell(r, 13).value),
                "carotenoid": _num(ws.cell(r, 14).value),
                # после хранения: Ca(21) сумм(26) Xk(27)
                "chl_sum_stored": _num(ws.cell(r, 26).value),
                "carotenoid_stored": _num(ws.cell(r, 27).value),
                # аскорбиновая к-та, мг/100 г (33)
                "vit_c": _num(ws.cell(r, 33).value),
                # сахара % свежее(36) / после хранения(39)
                "sugar": _num(ws.cell(r, 36).value),
                "sugar_stored": _num(ws.cell(r, 39).value),
                # нитраты (ед. лаборатории) свежее(42) / после хранения(45)
                "nitrate": _num(ws.cell(r, 42).value),
                "nitrate_stored": _num(ws.cell(r, 45).value),
            })
    return pd.DataFrame(rows)


# ---- Ассимиляционная площадь --------------------------------------------
def parse_assim():
    ws = WB["асс пл"]
    fac, cur = [], None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            fac.append({"cultivar": cur, "variant": int(var),
                        "assim_area_m2": _num(ws.cell(r, 8).value)})  # Среднее
    scr = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 13).value
        if isinstance(name, str) and name.strip() and name.strip() != "Сорт":
            vals = [_num(ws.cell(r, c).value) for c in range(14, 19)]
            vals = [v for v in vals if not np.isnan(v)]
            # отбрасываем явные опечатки (>0.1 м2 при типичных 0.02-0.05)
            vals = [v for v in vals if v < 0.1]
            if vals:
                scr.append({"cultivar": name.strip(), "assim_area_m2": np.mean(vals)})
    return pd.DataFrame(fac), pd.DataFrame(scr, columns=["cultivar", "assim_area_m2"])


# ---- Сухое вещество -----------------------------------------------------
def parse_drymatter():
    ws = WB["сух в-во"]
    # факторный: Сорт(1),Вариант(2),СВ%(7); строки могут иметь несколько навесок -> усредняем по variant
    fac, cur, curvar = [], None, None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if isinstance(var, (int, float)):
            curvar = int(var)
        sv = _num(ws.cell(r, 7).value)
        if cur in CORE and curvar and not np.isnan(sv):
            fac.append({"cultivar": cur, "variant": curvar, "dm_pct": sv})
    facdf = pd.DataFrame(fac).groupby(["cultivar", "variant"], as_index=False)["dm_pct"].mean()
    # скрининг: Сорт(11),СВ%(16)
    scr = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 11).value
        sv = _num(ws.cell(r, 16).value)
        if isinstance(name, str) and name.strip() and not np.isnan(sv):
            scr.append({"cultivar": name.strip(), "dm_pct": sv})
    return facdf, pd.DataFrame(scr, columns=["cultivar", "dm_pct"])


# ---- Водный дефицит / тургесцентность -----------------------------------
def parse_water_deficit():
    ws = WB["водный деф"]
    fac, cur = [], None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            fac.append({"cultivar": cur, "variant": int(var),
                        "water_deficit_pct": _num(ws.cell(r, 13).value),
                        "rel_turgor_pct": _num(ws.cell(r, 14).value)})
    scr = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 16).value
        if isinstance(name, str) and name.strip() and name not in ("сорт", "Сорт"):
            scr.append({"cultivar": name.strip(),
                        "water_deficit_pct": _num(ws.cell(r, 28).value),
                        "rel_turgor_pct": _num(ws.cell(r, 29).value)})
    return pd.DataFrame(fac), pd.DataFrame(scr)


# ---- Водоудерживающая способность ---------------------------------------
def parse_water_hold():
    ws = WB["водоудерж способ"]
    fac, cur = [], None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            fac.append({"cultivar": cur, "variant": int(var),
                        "water_hold_pct": _num(ws.cell(r, 15).value),
                        "mobile_water_pct": _num(ws.cell(r, 16).value)})
    scr = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 17).value
        if isinstance(name, str) and name.strip() and name not in ("сорт", "Сорт"):
            scr.append({"cultivar": name.strip(),
                        "water_hold_pct": _num(ws.cell(r, 26).value),
                        "mobile_water_pct": _num(ws.cell(r, 27).value)})
    return pd.DataFrame(fac), pd.DataFrame(
        scr, columns=["cultivar", "water_hold_pct", "mobile_water_pct"])


# ---- Хранение (потери массы) --------------------------------------------
def parse_storage():
    ws = WB["хранение"]
    fac, cur = [], None
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if isinstance(s, str) and s.strip():
            cur = s.strip()
        var = ws.cell(r, 2).value
        if cur in CORE and isinstance(var, (int, float)):
            fac.append({"cultivar": cur, "variant": int(var),
                        "mass_before_g": _num(ws.cell(r, 5).value),
                        "mass_after_g": _num(ws.cell(r, 7).value),
                        "loss_pct": _num(ws.cell(r, 9).value)})
    return pd.DataFrame(fac)


# ---- Фенология (всхожесть) ----------------------------------------------
SOWING = pd.Timestamp("2026-03-03")


def parse_phenology():
    ws = WB["фенология"]
    rows = []
    for r in range(3, 30):
        name = ws.cell(r, 1).value
        if not (isinstance(name, str) and name.strip()):
            continue
        single = ws.cell(r, 2).value   # единичные всходы (дата) или None
        mass = ws.cell(r, 3).value     # массовые всходы (дата) или примечание
        note = None
        if isinstance(mass, str):
            note = mass
            mass = None
        single_d = single if hasattr(single, "year") else None
        mass_d = mass if hasattr(mass, "year") else None
        rows.append({
            "cultivar": name.strip().capitalize(),
            "single_days": (pd.Timestamp(single_d) - SOWING).days if single_d else np.nan,
            "mass_days": (pd.Timestamp(mass_d) - SOWING).days if mass_d else np.nan,
            "note": note,
        })
    return pd.DataFrame(rows)


def main():
    spad_f, spad_s = parse_spad()
    bio = parse_biometry()
    chem = parse_biochem()
    assim_f, assim_s = parse_assim()
    dm_f, dm_s = parse_drymatter()
    wd_f, wd_s = parse_water_deficit()
    wh_f, wh_s = parse_water_hold()
    storage = parse_storage()
    phen = parse_phenology()

    # ---- факторная сводка ----
    def collapse(df):
        # одна строка на (сорт,вариант): усредняем числовые дубли ключей
        num = df.select_dtypes("number").columns.difference(["variant"])
        return df.groupby(["cultivar", "variant"], as_index=False)[list(num)].mean()

    fac = collapse(spad_f).merge(collapse(bio), on=["cultivar", "variant"], how="outer")
    for df in (chem, assim_f, dm_f, wd_f, wh_f, storage):
        fac = fac.merge(collapse(df), on=["cultivar", "variant"], how="outer")
    fac = fac.merge(LIGHT[["variant", "name", "recipe", "ppfd", "power_w", "dli"]],
                    on="variant", how="left")
    fac = fac.sort_values(["cultivar", "variant"]).reset_index(drop=True)

    # площадь яруса 1.785 м2; листьев на лоток считать не будем — нет плотности.
    # энергоэффективность: товарная масса 1 растения на Вт
    fac["mkt_per_w_g"] = fac["leaf_mass_mkt_g"] / fac["power_w"]

    # ---- сводка скрининга ----
    scr = spad_s.merge(assim_s, on="cultivar", how="outer")
    scr = scr.merge(dm_s, on="cultivar", how="outer")
    scr = scr.merge(wd_s, on="cultivar", how="outer")
    scr = scr.merge(wh_s, on="cultivar", how="outer")

    fac.to_csv(OUT / "factorial.csv", index=False)
    scr.to_csv(OUT / "screening.csv", index=False)
    storage.to_csv(OUT / "storage.csv", index=False)
    phen.to_csv(OUT / "phenology.csv", index=False)
    LIGHT.to_csv(OUT / "light_meta.csv", index=False)

    print("=== LIGHT ===")
    print(LIGHT.to_string(index=False))
    print("\n=== FACTORIAL (форма) ===", fac.shape)
    print(fac[["cultivar", "variant", "recipe", "spad_mean", "leaf_mass_mkt_g",
               "dm_pct", "vit_c", "sugar", "nitrate", "loss_pct"]].to_string(index=False))
    print("\n=== SCREENING ===", scr.shape)
    print(scr.to_string(index=False))
    print("\n=== PHENOLOGY ===")
    print(phen.to_string(index=False))


if __name__ == "__main__":
    main()
